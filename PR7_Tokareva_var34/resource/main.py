import os
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import load_workbook


class SimpleMarketingCampaignApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Campaign Form")
        self.root.geometry("760x620")
        self.root.resizable(False, False)

        self.var_campaign_input_unique = tk.StringVar()
        self.var_marketer_input_unique = tk.StringVar()
        self.var_channel_input_unique = tk.StringVar()
        self.var_priority_input_unique = tk.StringVar()

        self.record_id_counter_unique = 1

        self._build_ui_unique()
        self._bind_events_unique()
        self._update_add_button_state_unique()
        self.entry_campaign_unique.focus_set()

        # Автозаполнение из Excel для демонстрационного скриншота результата.
        self.root.after(300, self._autoload_from_excel_unique)

    def _build_ui_unique(self) -> None:
        self.frame_main_unique = ttk.Frame(self.root, padding=16)
        self.frame_main_unique.pack(fill="both", expand=True)

        self.label_title_unique = ttk.Label(
            self.frame_main_unique,
            text="Marketing Campaign Registration",
            font=("Segoe UI", 14, "bold")
        )
        self.label_title_unique.pack(anchor="w", pady=(0, 14))

        self.label_campaign_unique = ttk.Label(self.frame_main_unique, text="Campaign")
        self.label_campaign_unique.pack(anchor="w")
        self.entry_campaign_unique = ttk.Entry(
            self.frame_main_unique,
            textvariable=self.var_campaign_input_unique,
            name="entry_campaign_unique"
        )
        self.entry_campaign_unique.pack(fill="x", pady=(4, 10))

        self.label_marketer_unique = ttk.Label(self.frame_main_unique, text="Marketer")
        self.label_marketer_unique.pack(anchor="w")
        self.entry_marketer_unique = ttk.Entry(
            self.frame_main_unique,
            textvariable=self.var_marketer_input_unique,
            name="entry_marketer_unique"
        )
        self.entry_marketer_unique.pack(fill="x", pady=(4, 10))

        self.label_channel_unique = ttk.Label(self.frame_main_unique, text="Channel")
        self.label_channel_unique.pack(anchor="w")
        self.combobox_channel_unique = ttk.Combobox(
            self.frame_main_unique,
            textvariable=self.var_channel_input_unique,
            values=["Email", "Social Media", "Ads"],
            state="readonly",
            name="combobox_channel_unique"
        )
        self.combobox_channel_unique.pack(fill="x", pady=(4, 10))
        self.combobox_channel_unique.set("")

        self.label_priority_unique = ttk.Label(self.frame_main_unique, text="Priority")
        self.label_priority_unique.pack(anchor="w")
        self.combobox_priority_unique = ttk.Combobox(
            self.frame_main_unique,
            textvariable=self.var_priority_input_unique,
            values=["Low", "Medium", "High"],
            state="readonly",
            name="combobox_priority_unique"
        )
        self.combobox_priority_unique.pack(fill="x", pady=(4, 14))
        self.combobox_priority_unique.set("")

        self.frame_buttons_unique = ttk.Frame(self.frame_main_unique)
        self.frame_buttons_unique.pack(anchor="w", pady=(0, 14))

        self.button_add_unique = ttk.Button(
            self.frame_buttons_unique,
            text="Add",
            command=self._save_current_record_unique,
            state="disabled",
            name="button_add_unique"
        )
        self.button_add_unique.pack(side="left", padx=(0, 8))

        self.button_clear_unique = ttk.Button(
            self.frame_buttons_unique,
            text="Clear Form",
            command=self._clear_form_unique,
            name="button_clear_unique"
        )
        self.button_clear_unique.pack(side="left")

        self.label_output_title_unique = ttk.Label(
            self.frame_main_unique,
            text="Entered Data"
        )
        self.label_output_title_unique.pack(anchor="w")

        self.frame_table_unique = ttk.Frame(self.frame_main_unique)
        self.frame_table_unique.pack(fill="both", expand=True, pady=(6, 0))

        self.columns_unique = ("ID", "Campaign", "Marketer", "Channel", "Priority", "Status")
        self.tree_records_unique = ttk.Treeview(
            self.frame_table_unique,
            columns=self.columns_unique,
            show="headings",
            height=12,
            name="tree_records_unique"
        )

        for column_name_unique in self.columns_unique:
            self.tree_records_unique.heading(column_name_unique, text=column_name_unique)

        self.tree_records_unique.column("ID", width=50, anchor="center")
        self.tree_records_unique.column("Campaign", width=190, anchor="w")
        self.tree_records_unique.column("Marketer", width=160, anchor="w")
        self.tree_records_unique.column("Channel", width=110, anchor="center")
        self.tree_records_unique.column("Priority", width=90, anchor="center")
        self.tree_records_unique.column("Status", width=90, anchor="center")

        self.scrollbar_vertical_unique = ttk.Scrollbar(
            self.frame_table_unique,
            orient="vertical",
            command=self.tree_records_unique.yview
        )
        self.tree_records_unique.configure(yscrollcommand=self.scrollbar_vertical_unique.set)

        self.tree_records_unique.pack(side="left", fill="both", expand=True)
        self.scrollbar_vertical_unique.pack(side="right", fill="y")

    def _bind_events_unique(self) -> None:
        self.var_campaign_input_unique.trace_add("write", lambda *_: self._update_add_button_state_unique())
        self.var_marketer_input_unique.trace_add("write", lambda *_: self._update_add_button_state_unique())
        self.var_channel_input_unique.trace_add("write", lambda *_: self._update_add_button_state_unique())
        self.var_priority_input_unique.trace_add("write", lambda *_: self._update_add_button_state_unique())

        self.combobox_channel_unique.bind("<<ComboboxSelected>>", lambda _event: self._update_add_button_state_unique())
        self.combobox_priority_unique.bind("<<ComboboxSelected>>", lambda _event: self._update_add_button_state_unique())

    def _is_form_valid_unique(self) -> bool:
        return all([
            self.var_campaign_input_unique.get().strip(),
            self.var_marketer_input_unique.get().strip(),
            self.var_channel_input_unique.get().strip(),
            self.var_priority_input_unique.get().strip(),
        ])

    def _update_add_button_state_unique(self) -> None:
        self.button_add_unique.config(state="normal" if self._is_form_valid_unique() else "disabled")

    def _save_current_record_unique(self) -> None:
        if not self._is_form_valid_unique():
            messagebox.showwarning("Validation Error", "Fill in all fields.")
            return

        status_value_unique = self._build_status_unique(self.var_priority_input_unique.get().strip())
        self.tree_records_unique.insert(
            "",
            tk.END,
            values=(
                self.record_id_counter_unique,
                self.var_campaign_input_unique.get().strip(),
                self.var_marketer_input_unique.get().strip(),
                self.var_channel_input_unique.get().strip(),
                self.var_priority_input_unique.get().strip(),
                status_value_unique,
            ),
        )
        self.record_id_counter_unique += 1

    def _build_status_unique(self, priority_value_unique: str) -> str:
        status_map_unique = {
            "High": "Urgent",
            "Medium": "Planned",
            "Low": "Queued",
        }
        return status_map_unique.get(priority_value_unique, "New")

    def _clear_form_unique(self) -> None:
        self.var_campaign_input_unique.set("")
        self.var_marketer_input_unique.set("")
        self.var_channel_input_unique.set("")
        self.var_priority_input_unique.set("")
        self._update_add_button_state_unique()
        self.entry_campaign_unique.focus_set()

    def _set_form_values_unique(self, campaign_unique: str, marketer_unique: str, channel_unique: str, priority_unique: str) -> None:
        self.var_campaign_input_unique.set(campaign_unique)
        self.var_marketer_input_unique.set(marketer_unique)
        self.var_channel_input_unique.set(channel_unique)
        self.var_priority_input_unique.set(priority_unique)
        self._update_add_button_state_unique()

    def _resource_excel_path_unique(self) -> str:
        current_dir_unique = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(current_dir_unique, "file.xlsx")

    def _autoload_from_excel_unique(self) -> None:
        excel_path_unique = self._resource_excel_path_unique()
        if not os.path.exists(excel_path_unique):
            return

        try:
            workbook_unique = load_workbook(excel_path_unique)
            worksheet_unique = workbook_unique.active
        except Exception as error_unique:
            messagebox.showerror("Excel Error", f"Cannot open file.xlsx\n{error_unique}")
            return

        rows_unique = list(worksheet_unique.iter_rows(values_only=True))
        if not rows_unique:
            return

        header_unique = [str(cell).strip() if cell is not None else "" for cell in rows_unique[0]]
        data_rows_unique = rows_unique[1:]

        required_columns_unique = ["Campaign", "Marketer", "Channel", "Priority"]
        if any(column_unique not in header_unique for column_unique in required_columns_unique):
            messagebox.showerror(
                "Excel Error",
                "В файле file.xlsx должны быть колонки: Campaign, Marketer, Channel, Priority"
            )
            return

        column_index_map_unique = {name_unique: header_unique.index(name_unique) for name_unique in required_columns_unique}

        for row_unique in data_rows_unique:
            if row_unique is None:
                continue

            campaign_unique = str(row_unique[column_index_map_unique["Campaign"]] or "").strip()
            marketer_unique = str(row_unique[column_index_map_unique["Marketer"]] or "").strip()
            channel_unique = str(row_unique[column_index_map_unique["Channel"]] or "").strip()
            priority_unique = str(row_unique[column_index_map_unique["Priority"]] or "").strip()

            if not any([campaign_unique, marketer_unique, channel_unique, priority_unique]):
                continue

            self._set_form_values_unique(
                campaign_unique,
                marketer_unique,
                channel_unique,
                priority_unique,
            )
            self._save_current_record_unique()

        # Оставляем в форме последнюю запись для наглядного скриншота результата.
        children_unique = self.tree_records_unique.get_children()
        if children_unique:
            last_values_unique = self.tree_records_unique.item(children_unique[-1], "values")
            self._set_form_values_unique(
                str(last_values_unique[1]),
                str(last_values_unique[2]),
                str(last_values_unique[3]),
                str(last_values_unique[4]),
            )


def main() -> None:
    root_window_unique = tk.Tk()
    app_instance_unique = SimpleMarketingCampaignApp(root_window_unique)
    root_window_unique.mainloop()


if __name__ == "__main__":
    main()
